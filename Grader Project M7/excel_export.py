import matplotlib
matplotlib.use("Agg")   # 🔥 Required for EXE, Flask, server-safe image generation

"""
Excel export module for M7Pro Grader
Creates Excel file with:
- Sheet 1: Results DataFrame (all grades)
- Sheet 2: Plots (both charts as per wireframe)
"""
import io
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image


def make_excel_from_df(df, module, feedback):
    """
    Create Excel file with results in one sheet and plots in another sheet
    As per wireframe requirements
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: Results DataFrame
        df.to_excel(writer, sheet_name="results", index=False)

        # Sheet 2: Plots (TWO charts)
        wb = writer.book
        wb.create_sheet("plots")
        ws = wb["plots"]
        
        ws['A1'] = f"M7Pro Grader - {module} - Dashboard Visualizations"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        
        # CHART 1: Average criteria scores
        cols = [c for c in df.columns if c not in ["file", "total_grade"]]
        if cols:
            plt.figure(figsize=(7, 4))
            colors = ['#4CAF50', '#2196F3', '#FF9800', '#9C27B0', '#F44336']
            df[cols].mean().plot(kind="bar", color=colors[:len(cols)], edgecolor='black', linewidth=1.2)
            plt.title(f"{module} - Average Scores by Criteria", fontsize=14, fontweight='bold')
            plt.ylabel("Weighted Score (0-100)", fontsize=11)
            plt.xlabel("Criteria", fontsize=11)
            plt.xticks(rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3)
            plt.tight_layout()

            img1 = io.BytesIO()
            plt.savefig(img1, format="png", dpi=110)
            plt.close()
            img1.seek(0)

            ws.add_image(Image(img1), "A3")
        
        # CHART 2: Grade 1 distribution
        grade_1_count = (df["total_grade"] == 1).sum()
        other_count = len(df) - grade_1_count
        total_students = len(df)
        
        grade_1_pct = (grade_1_count / total_students * 100) if total_students > 0 else 0
        other_pct = (other_count / total_students * 100) if total_students > 0 else 0
        
        plt.figure(figsize=(7, 4))
        categories = ['Grade = 1\n(Failed)', 'Other Grades\n(Passed)']
        counts = [grade_1_count, other_count]
        colors_bar = ['#ef4444', '#10b981']
        
        bars = plt.bar(categories, counts, color=colors_bar, edgecolor='black', linewidth=1.2, alpha=0.8)
        
        for bar, count, pct in zip(bars, counts, [grade_1_pct, other_pct]):
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{count} students\n({pct:.1f}%)',
                    ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        plt.title(f"{module} - Students with Grade 1 vs Other Grades", fontsize=14, fontweight='bold')
        plt.ylabel("Number of Students", fontsize=11)
        plt.grid(axis='y', alpha=0.3)
        plt.tight_layout()

        img2 = io.BytesIO()
        plt.savefig(img2, format="png", dpi=110)
        plt.close()
        img2.seek(0)

        ws['A33'] = "📉 Bar Chart - Students with Grade 1"
        ws['A34'] = "Number and percentage of students who received a failing grade (1)"

        ws.add_image(Image(img2), "A36")

    output.seek(0)
    return output.getvalue()
